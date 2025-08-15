from fastapi.responses import JSONResponse
import traceback

from fastapi import FastAPI, Request, HTTPException, Depends, Header, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, RedirectResponse, JSONResponse
from pydantic import BaseModel
from typing import Optional
import io
from openpyxl import Workbook
from reportlab.pdfgen import canvas
import os
from app.engine import run_calc, Inputs
from app.equilibrium import solve_equilibrium_principal
from .calculator import compute

app = FastAPI(title="IPA Calculator API")

# CORS for local dev
origins = [o.strip() for o in os.getenv("FRONTEND_ORIGINS","http://localhost:3000,http://127.0.0.1:3000").split(",") if o.strip()]
koyeb = os.getenv("KOYEB_APP_DOMAIN")
if koyeb:
    origins.append(f"https://{koyeb}")
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class LoginIn(BaseModel):
    username: str
    password: str

@app.post("/auth/login/json")
def login_json(body: LoginIn):
    if body.username == "admin@example.com" and body.password == "admin":
        return {"access_token": "demo-token", "token_type": "bearer"}
    raise HTTPException(status_code=401, detail="Invalid credentials")

# very light auth dependency that accepts the demo token
def require_auth(authorization: Optional[str] = Header(None, alias="Authorization")):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing bearer token")
    token = authorization.split(" ", 1)[1].strip()
    if token != "demo-token":
        raise HTTPException(status_code=401, detail="Invalid token")
    return True

@app.get("/health")
def health():
    return {"status": "ok"}

@app.post("/calculate")
async def calculate(req: Request, _=Depends(require_auth)):
    body = await req.json()
    try:
        inp = Inputs(
            principal=float(body.get("principal", 0) or body.get("Inputs__principal", 0)),
            rate=float(body.get("rate", 0) or body.get("Inputs__rate", 0)),
            term_months=int(body.get("term_months", 0) or body.get("Inputs__term_months", 0)),
            balloon=float(body.get("balloon", 0) or body.get("Inputs__balloon", 0)),
            vat_rate=float(body.get("vat_rate", 0.18) or body.get("Inputs__vat_rate", 0.18)),
            asset_vat=float(body.get("asset_vat", 0) or body.get("Inputs__asset_vat", 0)),
            telematics_monthly=float(body.get("telematics_monthly", 0) or body.get("Inputs__telematics_monthly", 0)),
            include_irc=bool(body.get("include_irc", True) if body.get("include_irc", True) is not None else True),
            include_banking=bool(body.get("include_banking", True) if body.get("include_banking", True) is not None else True),
        )
    except Exception:
        # Fallback: if constructor signature differs, let engine parse dict directly
        inp = body

    res = run_calc(inp)
    # Ensure a totals dict is present for tests/clients
    if not isinstance(res.get("totals"), dict):
        t = {}
        for k in ("annuity","ipa_vat","asset_vat","vat_delta"):
            if k in res: t[k] = res[k]
        res["totals"] = t
    return res
@app.post("/export/xlsx")
async def export_xlsx(req: Request, _=Depends(require_auth)):
    data = await req.json()
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculation"
    ws["A1"] = "Example Metric"
    ws["B1"] = "Another"
    ws["A2"] = 123.45
    ws["B2"] = 67.89
    ws["A4"] = "Payload Echo"
    ws["B4"] = str(data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="calculation.xlsx"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )

@app.post("/export/pdf")
async def export_pdf(req: Request, _=Depends(require_auth)):
    data = await req.json()
    buf = io.BytesIO()
    c = canvas.Canvas(buf)
    c.setFont("Helvetica", 12)
    c.drawString(72, 800, "IPA Calculator - Demo PDF")
    c.drawString(72, 780, "Example Metric: 123.45")
    c.drawString(72, 760, "Another: 67.89")
    c.drawString(72, 740, f"Payload: {str(data)[:80]}...")
    c.showPage()
    c.save()
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="calculation.pdf"'}
    return StreamingResponse(buf, media_type="application/pdf", headers=headers)


# Public (dev-only) endpoint: compute without auth
@app.post("/calculate_public")
async def calculate_public(body: Inputs):
    import os
    if os.getenv("DEV_ALLOW_PUBLIC_COMPUTE") != "1":
        raise HTTPException(status_code=404, detail="Not found")
    return run_calc(body)



# Dev-only: equilibrium compute (hidden), guarded by DEV_ALLOW_PUBLIC_COMPUTE
from fastapi import Body
from fastapi import HTTPException
import os

@app.post("/calculate_public_equilibrium", include_in_schema=False)
async def calculate_public_equilibrium(body: dict = Body(...)):
    if os.getenv("DEV_ALLOW_PUBLIC_COMPUTE") != "1":
        raise HTTPException(status_code=404, detail="Not found")
    # Solve for principal so VAT(IPA) ~= VAT(asset)
    return solve_equilibrium_principal(body)


from fastapi import Body, HTTPException
import os

@app.post("/_debug/equilibrium_scan", include_in_schema=False)
async def _debug_equilibrium_scan(body: dict = Body(...)):
    if os.getenv("DEV_ALLOW_PUBLIC_COMPUTE") != "1":
        raise HTTPException(status_code=404, detail="Not found")
    from app.equilibrium import equilibrium_error_for_principal
    err = equilibrium_error_for_principal(body)
    C = float(body.get("principal", 0) or 0)
    grid = [max(0.01, k*C) for k in (0.1,0.3,0.7,1.0,1.3,1.7,2.5)]
    return {"grid":[{"principal": round(x,2), "err": round(err(x),4)} for x in grid]}


@app.post("/calculate_public_equilibrium2", include_in_schema=False)
async def calculate_public_equilibrium2(body: dict = Body(...)):
    if os.getenv("DEV_ALLOW_PUBLIC_COMPUTE") != "1":
        raise HTTPException(status_code=404, detail="Not found")
    try:
        from app.equilibrium import solve_equilibrium_principal
        return solve_equilibrium_principal(body)
    except Exception as e:
        return JSONResponse(status_code=500, content={"error":"equilibrium failed","message":str(e),"trace": traceback.format_exc()})


@app.post("/calculate_public_equilibrium_f", include_in_schema=False)
async def calculate_public_equilibrium_f(body: dict = Body(...)):
    if os.getenv("DEV_ALLOW_PUBLIC_COMPUTE") != "1":
        raise HTTPException(status_code=404, detail="Not found")
    from app.equilibrium import solve_equilibrium_f
    return solve_equilibrium_f(body)


@app.post("/calculate_public_equilibrium_fsolve", include_in_schema=False)
async def calculate_public_equilibrium_fsolve(body: dict = Body(...)):
    if os.getenv("DEV_ALLOW_PUBLIC_COMPUTE") != "1":
        raise HTTPException(status_code=404, detail="Not found")
    from app.equilibrium import solve_equilibrium_f_bisect
    return solve_equilibrium_f_bisect(body)


@app.post("/calculate_public_equilibrium_fsolve2", include_in_schema=False)
async def calculate_public_equilibrium_fsolve2(body: dict = Body(...)):
    if os.getenv("DEV_ALLOW_PUBLIC_COMPUTE") != "1":
        raise HTTPException(status_code=404, detail="Not found")
    try:
        from app.equilibrium import solve_equilibrium_f_bisect
        return solve_equilibrium_f_bisect(body)
    except Exception as e:
        return JSONResponse(status_code=500, content={
            "error": "fsolve failed",
            "message": str(e),
            "trace": traceback.format_exc(),
        })

